import requests
from bs4 import BeautifulSoup
import openpyxl
import re
import math
import os
import subprocess

URL_FIRST = '/list?sortType=votes'
URL_NOT_FIRST = '/list?sortType=USER_RATING&offset='
HEADERS = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.85 YaBrowser/21.11.1.932 Yowser/2.5 Safari/537.36',
               'accept': '*/*'}
FILE = 'results.xlsx'
HOST = 'https://readmanga.io'

def get_html(url):
    r = requests.get(url, headers=HEADERS)  #получение html-файла с сайта
    return r

def get_all_genres(soup): #создание единой строки с перечислением жанров для каждого тайтла
    block = soup.find('p', class_='elementList')
    genres = block.find_all('a', class_='element-link')
    all_genres = ''
    for genre in genres:
        all_genres += genre.get_text(strip=True) + ', '
    return all_genres[:len(all_genres)-2]

def get_title_content(html_title): #получение данных с конкретного тайтла
    soup = BeautifulSoup(html_title, 'html.parser')
    
    year = soup.find('span', class_='elem_year')
    if year: #проверка указан ли год
        year = int(year.find('a', class_='element-link').get_text(strip=True))
    else:
        year = 'Не указано'

    toms = re.findall("\d+", soup.find('div', class_='subject-meta').find('p').get_text())
    if toms: #проверка указаны ли тома
        toms = toms[0]
    else:
        toms = 0

    marks = 0
    all_right_blocks = soup.find('div', class_='rightContent').find_all('div', class_='rightBlock')
    if all_right_blocks:
        number = 0
        for block in range (0, len(all_right_blocks)):
            p = all_right_blocks[block].get_text().find('Количество закладок')
            if (not(p == -1)): #проверка указаны ли закладки
                number = block
        all_strongs = all_right_blocks[number].find_all('strong')
        for strong in all_strongs:
            marks += int((re.findall("\d+", strong.get_text()))[0])

    limitation = soup.find('span', class_='elem_limitation')
    if limitation: #проверка указано ли возрастное ограничение
        limitation = limitation.find('a', class_='element-link').get_text(strip=True)
    else:
        limitation = '0+'
    
    category = soup.find('span', class_='elem_category')
    if category: #проверка указана ли категория
        category = category.find('a', class_='element-link').get_text(strip=True)
    else:
        category = 'Не указано'

    discussions = soup.find('span', class_='badge').get_text(strip=True)
    if discussions: #проверка указаны ли обсуждения
        discussions = int(discussions)
    else:
        discussions = 0
    
    title = [{ #создание словаря с данными конкретного тайтла
        'title': soup.find('span', class_='name').get_text(strip=True),
        'genre': get_all_genres(soup),
        'category': category,
        'year': year,
        'toms': int(toms),
        'discussions': discussions,
        'marks': marks,
        'limitation': limitation,
    }]
    return title

def get_content(html, rows): #получение данных со всего сайта
    soup = BeautifulSoup(html, 'html.parser')
    items = soup.find_all('div', class_='tile')
    titles = []
    counter = 1
    while counter <= rows:
        #print (f'Обработка тайтла номер {counter} из {rows}')
        link = items[counter - 1].find('a', class_='non-hover').get('href'); #создание новой ссылки - для каждого тайтла
        html_title = get_html(HOST + link)
        if (html_title.status_code == 200):
            titles.extend(get_title_content(html_title.text)) #получение данных о текущем тайтле и запись их в общий словарь
        else:
            print ('Connection error!(')
        counter += 1
    return titles

def save_file(items, path): #сохранение результатов в excel-таблицу
    if (os.path.exists(FILE)):
        book = openpyxl.load_workbook(FILE)
        sheet = book.active
    else:
        book = openpyxl.Workbook()
        sheet = book.active
        sheet['A1'] = 'Название'
        sheet['B1'] = 'Категория'
        sheet['C1'] = 'Год выпуска'
        sheet['D1'] = 'Количество томов'
        sheet['E1'] = 'Обсуждения'
        sheet['F1'] = 'Жанры'
        sheet['G1'] = 'Количество закладок'
        sheet['H1'] = 'Возрастная рекомендация'

    row = sheet.max_row + 1
    for item in items:
        sheet[row][0].value = item['title']
        sheet[row][1].value = item['category']
        sheet[row][2].value = item['year']
        sheet[row][3].value = item['toms']
        sheet[row][4].value = item['discussions']
        sheet[row][5].value = item['genre']
        sheet[row][6].value = item['marks']
        sheet[row][7].value = item['limitation']
        row += 1
    
    book.save(path) #сохранение
    book.close() #закрытиеbook = openpyxl.Workbook()
    sheet = book.active

    sheet['A1'] = 'Название'
    sheet['B1'] = 'Категория'
    sheet['C1'] = 'Год выпуска'
    sheet['D1'] = 'Количество томов'
    sheet['E1'] = 'Обсуждения'
    sheet['F1'] = 'Жанры'
    sheet['G1'] = 'Количество закладок'
    sheet['H1'] = 'Возрастная рекомендация'

    row = 2
    for item in items:
        sheet[row][0].value = item['title']
        sheet[row][1].value = item['category']
        sheet[row][2].value = item['year']
        sheet[row][3].value = item['toms']
        sheet[row][4].value = item['discussions']
        sheet[row][5].value = item['genre']
        sheet[row][6].value = item['marks']
        sheet[row][7].value = item['limitation']
        row += 1
    
    book.save(path) #сохранение
    book.close() #закрытие

def get_pages_count(html): #подсчет количества страниц на сайте
    soup = BeautifulSoup(html, 'html.parser')
    pagination = soup.find_all('a', class_='step')
    if pagination: #проверка на наличие пагинации - перемещения между страницами
        return int(pagination[-1].get_text())
    else:
        return 1

    
def parse():
    URL = URL_FIRST
    print ('Установка соединения...')
    html = get_html(HOST + URL)
    if html.status_code == 200:
        num_of_titles = int(input('Введите необходимое количество тайтлов: '))
        titles = []
        all_pages_count = get_pages_count(html.text) #получение количества всех страниц на сайте
        if (all_pages_count * 70 < num_of_titles):
            print(f'Всего {all_pages_count * 70} тайтлов!')
            pages_count = all_pages_count
        else:
            pages_count = math.ceil(num_of_titles/70) #получение необходимого количества страниц 
        for page in range (1, pages_count + 1):
            if (page > 1):
                num = (page - 1) * 70
                URL = URL_NOT_FIRST + f'{num}' #составление новой ссылки для страниц, начиная со второй
            rows = 70
            if page == pages_count:
                rows = num_of_titles - 70 * (pages_count - 1)
                
            print (f'Парсинг страницы {page} из {pages_count}...')
            html = get_html(HOST + URL) #подключение к сайту
            titles.extend(get_content(html.text, rows)) #получение данных с сайта и добавление их в словарь

        print (f'Обработано {len(titles)} тайтлов')
        save_file(titles, FILE) #создание excel-файла и сохранение в него всех данных

        os.startfile(FILE) #запуск созданного excel-файла для Windows
        
        #subprocess.call(['open', FILE]) #запуск созданного excel-файла для Mac os
    else:
        print ('Server error(')

parse()
